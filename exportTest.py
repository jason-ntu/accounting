import io
from unittest.mock import patch
import sqlalchemy as sql
from export import ExportPage, ExportOption
from mock_db import MockDB
from accessor import ExecutionStatus as es
import const
import os
from openpyxl import load_workbook
from datetime import datetime

class TestExportPage(MockDB):

    @patch("sys.stdout", new_callable=io.StringIO)
    def test_show(self, _stdout):
        ExportPage.show()
        output_lines = _stdout.getvalue().strip().split('\n')
        self.assertEqual(output_lines[0], "%d: 選擇欲匯出的區間" % ExportOption.CHOOSE)
        self.assertEqual(output_lines[1], "%d: 回到上一頁" % ExportOption.BACK)

    @patch("sys.stdout", new_callable=io.StringIO)
    def test_hints(self, _stdout):
        hints = [(ExportPage.hint_start_date, "請輸入\"起始\"日期 (yyyy-mm-dd):\n"),
                 (ExportPage.hint_finish_date, "請輸入\"結束\"日期 (yyyy-mm-dd):\n"),
                 (ExportPage.hint_input_filename, "請輸入檔案名稱:\n")]
        for hint in hints:
            hint[0]()
            self.assertMultiLineEqual(_stdout.getvalue(), hint[1])
            _stdout.truncate(0)
            _stdout.seek(0)

    @patch("sys.stdout", new_callable=io.StringIO)
    @patch('builtins.input', side_effect=['0', '3', 'T', '1', '2'])
    def test_choose(self, _input,  _stdout):
        self.assertEqual(ExportPage.choose(), 1)
        self.assertEqual(_input.call_count, 4)
        self.assertEqual(_stdout.getvalue(), "請輸入 1 到 2 之間的數字:\n"*3)
        self.assertEqual(ExportPage.choose(), 2)

    @patch("sys.stdout", new_callable=io.StringIO)
    @patch.object(ExportPage, 'chooseInterval')
    def test_execute(self, _chooseInterval, _stdout):
        with self.mock_db_config:
            _chooseInterval.return_value = False
            ExportPage.execute(ExportOption.CHOOSE)
            self.assertEqual(_chooseInterval.call_count, 1)

            _chooseInterval.return_value = True
            ExportPage.execute(ExportOption.CHOOSE)
            self.assertEqual(_chooseInterval.call_count, 2)

            ExportPage.execute(ExportOption.BACK)
            self.assertEqual(_chooseInterval.call_count, 3)
        output_lines = _stdout.getvalue().strip().split('\n')
        self.assertEqual(output_lines[0], "%s操作失敗%s" %
                         (const.ANSI_RED, const.ANSI_RESET))
        self.assertEqual(output_lines[1], "%s操作成功%s" %
                         (const.ANSI_GREEN, const.ANSI_RESET))
        self.assertEqual(output_lines[2], "%s操作成功%s" %
                         (const.ANSI_GREEN, const.ANSI_RESET))

    @patch('sys.stdout', new_callable=io.StringIO)
    @patch('builtins.input', side_effect=['2022-01-10', '2023-01-10'])
    @patch.object(ExportPage, 'hint_finish_date')
    @patch.object(ExportPage, 'hint_start_date')
    @patch.object(ExportPage, 'inputFileName')
    def test_chooseInterval_with_valid_input(self, _inputFileName,  _hint_start_date, _hint_finish_date, _input, _stdout):
        ExportPage.chooseInterval()
        self.assertEqual(_stdout.getvalue().strip(), "起始日期 00:00:00 到結束日期 23:59:59")
        self.assertEqual(_input.call_count, 2)
        self.assertEqual(_inputFileName.call_count, 1)
        self.assertEqual(_hint_start_date.call_count, 1)
        self.assertEqual(_hint_finish_date.call_count, 1)

    @patch('sys.stdout', new_callable=io.StringIO)
    @patch('builtins.input', side_effect=['2022-02-30', '2023-0-01', '2022-01-01', '2022-13-10'])
    def test_chooseInterval_with_invalid_date(self, _input, _stdout):
        ExportPage.chooseInterval()
        self.assertIn('日期格式錯誤', _stdout.getvalue())
        _stdout.truncate(0)
        _stdout.seek(0)
        ExportPage.chooseInterval()
        self.assertEqual(_input.call_count, 4)
        self.assertIn('日期格式錯誤', _stdout.getvalue())

    @patch('sys.stdout', new_callable=io.StringIO)
    @patch('builtins.input', side_effect=['2022-01-10', '2022-01-09'])
    def test_chooseInterval_invaild_internal(self, _input, _stdout):
        ExportPage.chooseInterval()
        output_lines = _stdout.getvalue().strip().split('\n')
        self.assertEqual(output_lines[0], "起始日期 00:00:00 到結束日期 23:59:59")
        self.assertEqual(output_lines[1], "請輸入\"起始\"日期 (yyyy-mm-dd):")
        self.assertEqual(output_lines[2], "請輸入\"結束\"日期 (yyyy-mm-dd):")
        self.assertEqual(output_lines[3], "Error: 時間區間至少一天")

    @patch('builtins.input', side_effect=['filename'])
    @patch.object(ExportPage, 'hint_input_filename')
    @patch.object(ExportPage, 'exportFile')
    def testInputFileName(self, _exportFile, _hint_input_filename, _input):
        ExportPage.inputFileName("2023-05-01", "2023-05-30")
        self.assertEqual(_exportFile.call_count, 1)
        self.assertEqual(_hint_input_filename.call_count, 1)
        self.assertEqual(_input.call_count, 1)

    def testExportFile(self):
        with self.mock_db_config:
            ExportPage.setUp_connection_and_table()
            ExportPage.exportFile("2023-01-01", "2023-03-31", "testfile")
            ExportPage.tearDown_connection(es.NONE)
        current_files = os.listdir()
        self.assertIn("testfile.xlsx", current_files)
        workbook = load_workbook("testfile.xlsx")
        worksheet = workbook.active
        data_range = worksheet.iter_rows(min_row=2, values_only=True)
        record_0 = ('EXPENSE', '飲料', '現金', 101, 'comebuy', datetime(2023, 1, 1, 0, 0), datetime(2023, 1, 1, 0, 0), None, None)
        record_1 = ('EXPENSE', '食物', '現金', 87, '全家', datetime(2023, 2, 18, 0, 0), datetime(2023, 2, 18, 0, 0), None, None)
        record_2 = ('EXPENSE', '衣服', 'LinePay', 321, '百貨公司', datetime(2023, 3, 5, 0, 0), datetime(2023, 3, 5, 0, 0), None, '洋裝')
        record_3 = ('EXPENSE', '食物', '信用卡', 70, '百貨公司', datetime(2023, 3, 28, 0, 0), datetime(2023, 3, 30, 0, 0), None, 'coco')
        file_info = []
        for row in data_range:
            file_info.append(row)
        self.assertEqual(file_info[0], record_0)
        self.assertEqual(file_info[1], record_1)
        self.assertEqual(file_info[2], record_2)
        self.assertEqual(file_info[3], record_3)

        os.remove("testfile.xlsx")

    @patch.object(ExportPage, 'execute')
    @patch.object(ExportPage, 'choose', side_effect=[ExportOption.CHOOSE, ExportOption.BACK])
    @patch.object(ExportPage, 'show')
    def test_start(self, _show, _choose, _execute):
        ExportPage.start()
        self.assertEqual(_show.call_count, 2)
        self.assertEqual(_choose.call_count, 2)
        self.assertEqual(_execute.call_count, 1)